"""
Agent 5 — Renderer Agent.

Emits:
  - canonical_organogram.json
  - organogram_long_form.xlsx  (one row per person + reporting_chain string)
  - organogram.mermaid          (Levels 1–4 only, per #7)
  - organogram_kg.json          (Neo4j-import-ready)
  - organogram_geographic.json
  - organogram_legal_entity.json

No LLM calls. Pure I/O.
"""
from __future__ import annotations
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..schemas.types import CanonicalOrganogram, CanonicalNode


NAVY = "1F2A44"
TEAL = "0E7A75"
LIGHT = "F2F4F8"
WHITE = "FFFFFF"


class RendererAgent:
    """Agent 5 — emits canonical JSON and all derived formats."""

    def __init__(self, organogram: CanonicalOrganogram, output_dir: str | Path):
        self.org = organogram
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self._node_by_id: dict[str, CanonicalNode] = {n.id: n for n in self.org.nodes}

    # ------------------------------------------------------------------
    # PUBLIC ENTRY
    # ------------------------------------------------------------------
    def render_all(self) -> dict[str, Path]:
        """Render every format. Returns {format_name: path}."""
        paths = {}
        paths["canonical_json"] = self._render_canonical_json()
        paths["long_form_xlsx"] = self._render_long_form_xlsx()
        paths["mermaid"] = self._render_mermaid()
        paths["kg_json"] = self._render_kg_json()
        paths["geographic_json"] = self._render_geographic_json()
        paths["legal_entity_json"] = self._render_legal_entity_json()
        return paths

    # ------------------------------------------------------------------
    # 1) CANONICAL JSON
    # ------------------------------------------------------------------
    def _render_canonical_json(self) -> Path:
        path = self.output_dir / "canonical_organogram.json"
        path.write_text(
            json.dumps(self.org.to_dict(), indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return path

    # ------------------------------------------------------------------
    # 2) EXCEL LONG-FORM — option (b): level + reports_to_name + reporting_chain
    # ------------------------------------------------------------------
    def _build_chain(self, node: CanonicalNode) -> str:
        chain = [node.title_en]
        cur = node
        while cur.reports_to_id and cur.reports_to_id in self._node_by_id:
            cur = self._node_by_id[cur.reports_to_id]
            chain.append(f"{cur.name} ({cur.title_en})")
        return " > ".join(reversed(chain))

    def _render_long_form_xlsx(self) -> Path:
        path = self.output_dir / "organogram_long_form.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "People"

        headers = [
            "ID", "Name", "Title (Native)", "Title (English)", "Function",
            "Level", "Region", "Country", "Legal Entity",
            "Reports To (ID)", "Reports To (Name)", "Reporting Chain",
            "Source URL", "Source Type", "Inference Note",
        ]
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(top=thin, bottom=thin, left=thin, right=thin)
        header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
        header_fill = PatternFill("solid", fgColor=NAVY)
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = border

        # Sort nodes: by level, then by function, then by name
        nodes_sorted = sorted(
            self.org.nodes,
            key=lambda n: (n.level, n.function, n.name),
        )
        data_font = Font(name="Arial", size=10)
        alt_fill = PatternFill("solid", fgColor=LIGHT)

        for i, n in enumerate(nodes_sorted, start=2):
            reports_to_name = ""
            if n.reports_to_id and n.reports_to_id in self._node_by_id:
                reports_to_name = self._node_by_id[n.reports_to_id].name

            row = [
                n.id, n.name, n.title_native, n.title_en, n.function,
                n.level, n.region, n.country or "", n.legal_entity or "",
                n.reports_to_id or "", reports_to_name,
                self._build_chain(n),
                n.source or "", n.source_type or "", n.inference_note or "",
            ]
            for col, val in enumerate(row, start=1):
                c = ws.cell(row=i, column=col, value=val)
                c.font = data_font
                c.alignment = center if col == 6 else left
                c.border = border
                if i % 2 == 0:
                    c.fill = alt_fill

        widths = [12, 24, 22, 30, 18, 7, 14, 14, 22, 12, 24, 70, 50, 14, 36]
        for col, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(nodes_sorted)+1}"

        wb.save(path)
        return path

    # ------------------------------------------------------------------
    # 3) MERMAID — Levels 1–4 only (#7 = a)
    # ------------------------------------------------------------------
    def _safe_mermaid_id(self, node_id: str) -> str:
        # Mermaid IDs must be alphanumeric or underscore
        return node_id.replace("-", "_")

    def _mermaid_label(self, n: CanonicalNode) -> str:
        # Escape double-quotes inside the label
        name = n.name.replace('"', "'")
        title = n.title_en.replace('"', "'")
        return f'"{name}<br/>{title} (L{n.level})"'

    def _render_mermaid(self) -> Path:
        path = self.output_dir / "organogram.mermaid"
        nodes_in = [n for n in self.org.nodes if n.level <= 4]
        ids_in = {n.id for n in nodes_in}

        lines = ["graph TD"]
        # Declare nodes
        for n in nodes_in:
            mid = self._safe_mermaid_id(n.id)
            lines.append(f"    {mid}[{self._mermaid_label(n)}]")
        # Edges
        for n in nodes_in:
            if n.reports_to_id and n.reports_to_id in ids_in:
                lines.append(f"    {self._safe_mermaid_id(n.reports_to_id)} --> {self._safe_mermaid_id(n.id)}")

        # Add a footer note that levels 5+ are collapsed
        levels_below = [n for n in self.org.nodes if n.level > 4]
        if levels_below:
            lines.append("")
            lines.append(f"    %% {len(levels_below)} additional people exist at Levels 5–10. View them in the Excel long-form.")

        path.write_text("\n".join(lines), encoding="utf-8")
        return path

    # ------------------------------------------------------------------
    # 4) KNOWLEDGE GRAPH JSON  (Neo4j-import-ready)
    # ------------------------------------------------------------------
    def _render_kg_json(self) -> Path:
        path = self.output_dir / "organogram_kg.json"

        # Node labels
        kg_nodes = []
        functions_seen = set()
        regions_seen = set()
        entities_seen = set()

        for n in self.org.nodes:
            kg_nodes.append({
                "label": "Person",
                "id": n.id,
                "properties": {
                    "name": n.name,
                    "title_en": n.title_en,
                    "title_native": n.title_native,
                    "level": n.level,
                    "source": n.source,
                },
            })
            functions_seen.add(n.function)
            if n.region:
                regions_seen.add(n.region)
            if n.legal_entity:
                entities_seen.add(n.legal_entity)

        for f in sorted(functions_seen):
            kg_nodes.append({"label": "Function", "id": f"fn::{f}",
                             "properties": {"name": f}})
        for r in sorted(regions_seen):
            kg_nodes.append({"label": "Region", "id": f"rg::{r}",
                             "properties": {"name": r}})
        for e in sorted(entities_seen):
            kg_nodes.append({"label": "LegalEntity", "id": f"le::{e}",
                             "properties": {"name": e}})

        # Edges
        kg_edges = []
        for n in self.org.nodes:
            if n.reports_to_id:
                kg_edges.append({"type": "REPORTS_TO", "from": n.id, "to": n.reports_to_id})
            kg_edges.append({"type": "BELONGS_TO_FUNCTION", "from": n.id, "to": f"fn::{n.function}"})
            if n.region:
                kg_edges.append({"type": "LOCATED_IN", "from": n.id, "to": f"rg::{n.region}"})
            if n.legal_entity:
                kg_edges.append({"type": "EMPLOYED_BY", "from": n.id, "to": f"le::{n.legal_entity}"})

        path.write_text(
            json.dumps({"nodes": kg_nodes, "edges": kg_edges}, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return path

    # ------------------------------------------------------------------
    # 5) GEOGRAPHIC PIVOT JSON
    # ------------------------------------------------------------------
    def _render_geographic_json(self) -> Path:
        path = self.output_dir / "organogram_geographic.json"
        by_country: dict[str, list[str]] = {}
        for n in self.org.nodes:
            key = n.country or "Unknown"
            by_country.setdefault(key, []).append(n.id)
        path.write_text(
            json.dumps({"by_country": by_country}, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return path

    # ------------------------------------------------------------------
    # 6) LEGAL-ENTITY GRAPH JSON
    # ------------------------------------------------------------------
    def _render_legal_entity_json(self) -> Path:
        path = self.output_dir / "organogram_legal_entity.json"
        path.write_text(
            json.dumps(self.org.legal_entity_graph, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        return path
