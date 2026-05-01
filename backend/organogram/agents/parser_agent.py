"""
Agent 1 — Parser & Ingestion Agent.

Reads CSV / Excel / JSON inputs and emits a list of PersonRecord.
Does fuzzy header matching for common synonyms.

Supports the vendor database format (51 columns including JOB_FUNCTION,
JOB_LEVEL, JOB_LOCATION_*, etc.) — these vendor fields are passed through
to the NLP Agent which trusts them first and falls back to rules.
"""
from __future__ import annotations
import csv
import json
from pathlib import Path
from typing import Optional

from ..schemas.types import PersonRecord


# Multi-key synonym map. Header lookup is case-insensitive.
HEADER_SYNONYMS = {
    "name":            {"name", "full name", "full_name", "employee name", "person name"},
    "first_name":      {"first_name", "first name", "given name"},
    "last_name":       {"last_name", "last name", "surname", "family name"},
    "title":           {"title", "designation", "job title", "job_title", "role", "position"},
    "company":         {"company", "company_name", "employer", "current company",
                        "organization", "org"},
    "source_url":      {"source_url", "linkedin", "linkedin url", "linkedin_url",
                        "profile url", "url"},
    "department":      {"department", "function", "team"},
    "geography":       {"geography", "office_location", "location"},
    "tenure":          {"tenure", "start_date", "joined", "since"},
    "reports_to_name": {"reports_to_name", "reports to", "manager", "supervisor"},
    "subsidiary":      {"subsidiary", "legal_entity_name", "legal entity", "entity"},

    # Vendor-database-specific fields
    "vendor_function": {"job_function"},
    "vendor_level":    {"job_level"},
    "vendor_persona":  {"persona"},

    "job_country":         {"job_location_country"},
    "job_country_code":    {"job_location_country_code"},
    "job_country_region":  {"job_location_country_region"},
    "job_continent":       {"job_location_continent"},
    "job_state":           {"job_location_state", "job_location_state_code"},
    "job_city":            {"job_location_city"},

    "job_org_linkedin_url": {"job_org_linkedin_url"},
    "email_domain":         {"email_domain"},
    "linkedin_industry":    {"linkedin_industry"},
    "linkedin_headline":    {"linkedin_headline"},

    # Fallback geography fields if job-specific aren't populated
    "fallback_country":      {"country_name"},
    "fallback_country_code": {"country_code"},
    "fallback_country_region": {"country_region"},
    "fallback_continent":    {"continent"},
    "fallback_state":        {"state_name", "state_code"},
    "fallback_city":         {"city"},
}


def _norm(s: str) -> str:
    return (s or "").lower().strip().replace("-", "_").replace(" ", "_")


def _map_headers(headers: list[str]) -> dict[str, str]:
    """Return {canonical_field: actual_header_in_input}."""
    mapping = {}
    norm_headers = {_norm(h): h for h in headers}
    for canonical, synonyms in HEADER_SYNONYMS.items():
        for syn in synonyms:
            n = _norm(syn)
            if n in norm_headers:
                mapping[canonical] = norm_headers[n]
                break
    return mapping


class ParserAgent:
    """Agent 1 — input ingestion."""

    def parse(self, input_path: str | Path) -> list[PersonRecord]:
        p = Path(input_path)
        suffix = p.suffix.lower()
        if suffix == ".csv":
            return self._parse_csv(p)
        if suffix in (".xlsx", ".xlsm"):
            return self._parse_xlsx(p)
        if suffix == ".json":
            return self._parse_json(p)
        if suffix in (".tsv", ".txt"):
            return self._parse_csv(p, delimiter="\t")
        raise ValueError(f"Unsupported input format: {suffix}")

    def _parse_csv(self, path: Path, delimiter: str = ",") -> list[PersonRecord]:
        with path.open(encoding="utf-8") as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            headers = reader.fieldnames or []
            mapping = _map_headers(headers)
            records = []
            for row in reader:
                rec = self._row_to_record(row, mapping)
                if rec.name and rec.title:  # skip empty rows
                    records.append(rec)
        return self._dedupe(records)

    def _parse_xlsx(self, path: Path) -> list[PersonRecord]:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h) if h is not None else "" for h in next(rows_iter)]
        mapping = _map_headers(headers)
        records = []
        for raw in rows_iter:
            row = {headers[i]: (raw[i] if raw[i] is not None else "")
                   for i in range(len(headers))}
            rec = self._row_to_record(row, mapping)
            if rec.name and rec.title:
                records.append(rec)
        return self._dedupe(records)

    def _parse_json(self, path: Path) -> list[PersonRecord]:
        data = json.loads(path.read_text(encoding="utf-8"))
        if not isinstance(data, list):
            raise ValueError("JSON input must be a flat list of records.")
        if not data:
            return []
        keys = list(data[0].keys())
        mapping = _map_headers(keys)
        records = []
        for row in data:
            rec = self._row_to_record(row, mapping)
            if rec.name and rec.title:
                records.append(rec)
        return self._dedupe(records)

    def _row_to_record(self, row: dict, mapping: dict) -> PersonRecord:
        def get(field: str) -> Optional[str]:
            header = mapping.get(field)
            if header is None:
                return None
            v = row.get(header)
            if v is None:
                return None
            s = str(v).strip()
            return s or None

        # Compose name: prefer FULL_NAME, else first+last
        name = get("name")
        if not name:
            first = get("first_name") or ""
            last = get("last_name") or ""
            name = f"{first} {last}".strip()

        # Geography: prefer JOB_LOCATION_*, fall back to person-level country
        job_country = get("job_country")
        job_country_code = get("job_country_code")
        if not job_country:
            job_country = get("fallback_country")
        if not job_country_code:
            job_country_code = get("fallback_country_code")

        job_state = get("job_state") or get("fallback_state")
        job_city = get("job_city") or get("fallback_city")
        job_continent = get("job_continent") or get("fallback_continent")
        job_country_region = get("job_country_region") or get("fallback_country_region")

        # Use the most authoritative geography signal for the legacy `geography` field
        geography = job_country or get("geography")

        # COMPANY_NAME may be blank — fall back to org LinkedIn URL or email domain
        company = get("company")
        if not company:
            org_url = get("job_org_linkedin_url")
            email_dom = get("email_domain")
            if org_url:
                # Extract slug from /company/<slug> URL
                slug = org_url.rstrip("/").split("/")[-1]
                company = slug.replace("-", " ").title()
            elif email_dom:
                company = email_dom.split(".")[0].title()
            else:
                company = ""

        return PersonRecord(
            name=name or "",
            title=get("title") or "",
            company=company,
            source_url=get("source_url") or "",
            department=get("department"),
            geography=geography,
            tenure=get("tenure"),
            reports_to_name=get("reports_to_name"),
            subsidiary=get("subsidiary"),
            # Vendor classification (passed through to NLP Agent)
            vendor_function=get("vendor_function"),
            vendor_level=get("vendor_level"),
            vendor_persona=get("vendor_persona"),
            # Rich location
            job_country=job_country,
            job_country_code=job_country_code,
            job_country_region=job_country_region,
            job_continent=job_continent,
            job_state=job_state,
            job_city=job_city,
            # Fallback identifiers
            job_org_linkedin_url=get("job_org_linkedin_url"),
            email_domain=get("email_domain"),
            linkedin_industry=get("linkedin_industry"),
            linkedin_headline=get("linkedin_headline"),
        )

    def _dedupe(self, records: list[PersonRecord]) -> list[PersonRecord]:
        seen = set()
        out = []
        for r in records:
            key = (r.name.lower(), r.company.lower(), r.source_url.lower())
            if key in seen:
                continue
            seen.add(key)
            out.append(r)
        return out
