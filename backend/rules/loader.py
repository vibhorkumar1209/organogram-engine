"""
Organogram Engine — Runtime Excel Rule Loader
==============================================
Reads both hierarchy Excel files once at import time.
All rule consumers (classifier.py, structural_engine.py) import from here
so the Excel files are the single source of truth — no external file references needed.

Files loaded:
  rules/Global_Designation_Hierarchy.xlsx  →  TITLE_TO_GRADE
  rules/Global_Org_Hierarchy.xlsx          →  CANONICAL_L0_DEPTS, L0_DEPT_SUBS
"""
from __future__ import annotations

import logging
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

_LOG = logging.getLogger(__name__)

_RULES_DIR   = Path(__file__).parent
_DESIG_FILE  = _RULES_DIR / "Global_Designation_Hierarchy.xlsx"
_ORG_FILE    = _RULES_DIR / "Global_Org_Hierarchy.xlsx"

# Pre-compiled helpers
_GRADE_RE  = re.compile(r"^G(\d+)", re.IGNORECASE)
_PAREN_RE  = re.compile(r"\s*\([^)]*\)")          # strip "(CEO)" notes
_SLASH_RE  = re.compile(r"\s*/\s*")               # split "Chairman / Chairperson"
_ACRONYM_RE = re.compile(r"\(([A-Z]{2,8})\)")     # extract bare acronym from parens


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers
# ─────────────────────────────────────────────────────────────────────────────

def _norm(raw: str) -> str:
    """Lowercase, strip parenthetical notes, collapse whitespace."""
    return _PAREN_RE.sub("", raw).strip().lower()


def _load_designation_hierarchy() -> Dict[str, int]:
    """
    Parse Master Designations sheet → {normalized_title: grade_int}.

    For every row with a valid grade (G0–G11):
      - Primary Title             → added as-is (normalised)
      - Alternate Titles / Variants (comma-separated) → each added
      - Slash-separated variants ("Chairman / Chairperson") → each part added
      - Parenthetical acronym extraction: "CEO" from "(CEO)" → added as bare token
    """
    if not _DESIG_FILE.exists():
        _LOG.warning("Designation hierarchy not found: %s", _DESIG_FILE)
        return {}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(_DESIG_FILE), read_only=True, data_only=True)
        ws = wb["Master Designations"]

        title_to_grade: Dict[str, int] = {}

        for row in ws.iter_rows(values_only=True):
            if not row or len(row) < 4:
                continue
            # Columns: Grade Level | Primary Department | Function | Primary Title | Alternate Titles …
            grade_cell    = row[0]
            primary_title = row[3]
            alt_titles    = row[4] if len(row) > 4 else None

            if not grade_cell or not primary_title:
                continue
            m = _GRADE_RE.match(str(grade_cell).strip())
            if not m:
                continue
            grade = int(m.group(1))

            # Collect all raw title strings for this row
            raw_list: List[str] = [str(primary_title)]
            if alt_titles:
                raw_list.extend(str(alt_titles).split(","))

            for raw in raw_list:
                raw = raw.strip()
                if not raw or raw.lower() in ("none", ""):
                    continue

                # 1. Full normalised form
                key = _norm(raw)
                if key and key not in title_to_grade:
                    title_to_grade[key] = grade

                # 2. Slash-split variants ("Chairman / Chairperson of the Board")
                for part in _SLASH_RE.split(raw):
                    part_key = _norm(part)
                    if part_key and part_key not in title_to_grade:
                        title_to_grade[part_key] = grade

                # 3. Bare acronym from parentheses: "(CEO)" → "ceo"
                for acronym in _ACRONYM_RE.findall(raw):
                    acr_key = acronym.lower()
                    if acr_key not in title_to_grade:
                        title_to_grade[acr_key] = grade

        wb.close()
        _LOG.info(
            "Loaded %d title→grade mappings from %s",
            len(title_to_grade), _DESIG_FILE.name,
        )
        return title_to_grade

    except Exception as exc:
        _LOG.error("Failed to load %s: %s", _DESIG_FILE.name, exc)
        return {}


def _load_org_hierarchy() -> Tuple[List[str], Dict[str, List[str]]]:
    """
    Parse Master Hierarchy sheet.

    Returns:
      l0_depts — ordered list of canonical L0 department names
      l0_subs  — {l0_dept: [l1_sub_dept, …]} for downstream dept mapping
    """
    if not _ORG_FILE.exists():
        _LOG.warning("Org hierarchy not found: %s", _ORG_FILE)
        return [], {}

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(_ORG_FILE), read_only=True, data_only=True)
        ws = wb["Master Hierarchy"]

        l0_depts: List[str] = []
        l0_subs:  Dict[str, List[str]] = {}
        current_l0: Optional[str] = None

        # Column layout: Level | L0 Name | L1 Name | L2 | L3 | L4/Notes
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            level = str(row[0]).strip()

            if level == "L0":
                # L0 department name sits in column index 1
                name = row[1]
                if name:
                    current_l0 = str(name).strip()
                    if current_l0 not in l0_subs:
                        l0_depts.append(current_l0)
                        l0_subs[current_l0] = []

            elif level == "L1" and current_l0:
                # L1 sub-department name sits in column index 2
                sub = row[2]
                if sub:
                    l0_subs[current_l0].append(str(sub).strip())

        wb.close()
        _LOG.info(
            "Loaded %d L0 departments from %s",
            len(l0_depts), _ORG_FILE.name,
        )
        return l0_depts, l0_subs

    except Exception as exc:
        _LOG.error("Failed to load %s: %s", _ORG_FILE.name, exc)
        return [], {}


# ─────────────────────────────────────────────────────────────────────────────
# Public exports  (populated once at import time)
# ─────────────────────────────────────────────────────────────────────────────

#: {normalized_title: grade_int} — sourced from Global_Designation_Hierarchy.xlsx
TITLE_TO_GRADE: Dict[str, int] = _load_designation_hierarchy()

_l0_list, _l0_subs = _load_org_hierarchy()

#: Ordered canonical L0 department names from Global_Org_Hierarchy.xlsx
CANONICAL_L0_DEPTS: List[str] = _l0_list

#: {l0_dept: [l1_sub_dept, …]} sub-department index
L0_DEPT_SUBS: Dict[str, List[str]] = _l0_subs
